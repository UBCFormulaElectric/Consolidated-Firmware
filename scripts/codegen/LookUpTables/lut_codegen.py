from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os

number_of_voltages = 3
number_of_columns = 200
number_of_rows = 200

def write_sheet_to_file(workbook, file, comma_termination = True):
    worksheet = workbook.active

    file.write(
        '\t{\n'
    )

    for row in range(1,number_of_rows+1):
        file.write(
            '\t\t{'
        )
        for col in range(1,number_of_columns+1):
            char = get_column_letter(col)
            val = int(worksheet[char + str(row)].value*100)

            if col == 1:
                file.write(
                    str(val)
                )
            else:
                file.write(
                    ',\t ' + str(val)
                )
        if row == number_of_rows:
            file.write(
                '}\n'
            )
        else:
            file.write(
                '},\n'
            )
    if comma_termination:
        file.write(
            '\t},\n'
        )
    else:
        file.write(
            '\t}\n'
        )

#400V data
id_400v =  load_workbook('MotorData/id_400V_rebased.xlsx')
iq_400v =  load_workbook('MotorData/iq_400V_rebased.xlsx')
is_400v =  load_workbook('MotorData/is_400V_rebased.xlsx')

#500V data
id_500v =  load_workbook('MotorData/id_500V_rebased.xlsx')
iq_500v =  load_workbook('MotorData/iq_500V_rebased.xlsx')
is_500v =  load_workbook('MotorData/is_500V_rebased.xlsx')

#600V data
id_600v =  load_workbook('MotorData/id_600V_rebased.xlsx')
iq_600v =  load_workbook('MotorData/iq_600V_rebased.xlsx')
is_600v =  load_workbook('MotorData/is_600V_rebased.xlsx')

base_path = os.getenv('PYTHONPATH')

#generate header (.h) file containing lookup table definition
with open(base_path + "/boards/INV/Inc/App/lookup_tables/id_peak_lut.h", 'w') as hdr_file:
    hdr_file.write(
        '#pragma once\n'
        '\n'
        '#include \"App_MotorLutInterface.h\"\n'
        '\n'
        'static const int16_t id_peak_lut_times100[LUT_NUM_VOLTAGES][LUT_NUM_ROWS][LUT_NUM_COLUMNS] =\n'
        '{\n'
    )
    write_sheet_to_file(id_400v, hdr_file)
    write_sheet_to_file(id_500v, hdr_file)
    write_sheet_to_file(id_600v, hdr_file, comma_termination = False)
    hdr_file.write(
        '};\n'
    )

with open(base_path + "/boards/INV/Inc/App/lookup_tables/iq_peak_lut.h", 'w') as hdr_file:
    hdr_file.write(
            '#pragma once\n'
            '\n'
            '#include \"App_MotorLutInterface.h\"\n'
            '\n'
            'static const int16_t iq_peak_lut_times100[LUT_NUM_VOLTAGES][LUT_NUM_ROWS][LUT_NUM_COLUMNS] =\n'
            '{\n'
    )
    write_sheet_to_file(iq_400v, hdr_file)
    write_sheet_to_file(iq_500v, hdr_file)
    write_sheet_to_file(iq_600v, hdr_file, comma_termination = False)
    hdr_file.write(
        '};\n'
    )

with open(base_path + "/boards/INV/Inc/App/lookup_tables/is_peak_lut.h", 'w') as hdr_file:
    hdr_file.write(
            '#pragma once\n'
            '\n'
            '#include \"App_MotorLutInterface.h\"\n'
            '\n'
            'static const int16_t is_peak_lut_times100[LUT_NUM_VOLTAGES][LUT_NUM_ROWS][LUT_NUM_COLUMNS] =\n'
            '{\n'
    )
    write_sheet_to_file(is_400v, hdr_file)
    write_sheet_to_file(is_500v, hdr_file)
    write_sheet_to_file(is_600v, hdr_file, comma_termination = False)
    hdr_file.write(
        '};\n'
    )
